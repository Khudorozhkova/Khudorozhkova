import csv
import math
import re
import sys
from collections import Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
import doctest

'''
    Словарь для перевода валют
'''
currency_to_rub = {
    "AZN": 35,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}

class Vacancy:
    '''
        Класс для представления вакансий

        Attributes:
            dic (dict) словарь вакансий из csv файла
            salary (int) зп из csv файла
            is_needed (string) нужные вакансии из csv файла
        :return: void
    '''
    def __init__(self, dic: dict):
        self.dic = dic
        self.salary = Salary(dic)
        self.dic["year"] = int(dic["published_at"][:4])
        self.is_needed = dic["is_needed"]


class Salary:
    '''
        Класс для представленя зарплаты
        Attributes:
            salary_from (float) нижняя граница вилки оклада
            salary_to (float) верхняя граница вилки оклада
            salary_currency (string) Валюта
            averageSalary (int) средняя зарплата
            salaryRub (int) средняя зп в рублях
            >>> type(Salary(dic={"salary_from" : 100, "salary_to" : 200, "salary_currency" : "RUR"})).__name__
            'Salary'
            >>> Salary(dic={"salary_from" : 100, "salary_to" : 200, "salary_currency" : "RUR"}).averageSalary
            150.0
            >>> Salary(dic={"salary_from" : 100, "salary_to" : 200, "salary_currency" : "EUR"}).salaryRub
            8985.0
        :return: void
    '''
    def __init__(self, dic):
        self.salary_from = math.floor(float(dic["salary_from"]))
        self.salary_to = math.floor(float(dic["salary_to"]))
        self.salary_currency = dic["salary_currency"]
        self.averageSalary = (self.salary_to + self.salary_from) / 2
        self.salaryRub = currency_to_rub[self.salary_currency] * self.averageSalary


class DataSet:
    '''
        Класс для обработки данных из фалйа csv

        Attributes:
            inputValues (InputConect) входные данные
        :return: void
    '''

    def __init__(self):
        self.inputValues = InputConect()
        self.csv_reader()
        self.csv_filter()
        self.getYears()
        self.numberGraph()
        # self.printGraph()

    def csv_reader(self):
        '''
            Считывает информацию из файла в переменные
            :return: void
        '''
        with open(self.inputValues.fileName, "r", encoding='utf-8-sig', newline='') as csv_file:
            file = csv.reader(csv_file)
            self.firstLine = next(file)
            self.otherLines = [line for line in file
                               if not ("" in line) and len(line) == len(self.firstLine)] # a lot eat

    def tryToAdd(self, dic: dict, key, val) -> dict:
        '''
        Увеличние значения по ключу, если значение есть, если нет, то добавление записи в словарь
            :param dic: Входной словарь
            :param key: Ключ, которому добовляем значение
            :param val: Значение
            :return: dict
            >>> DataSet().tryToAdd(dic={'1' : 1}, key='1', val=1)
            {'1': 2}
            >>> DataSet().tryToAdd(dic={'1' : 1}, key='2', val=2)
            {'1': 1, '2': 2}
        '''
        try:
            dic[key] += val
        except:
            dic[key] = val
        return dic

    def csv_filter(self):
        '''
        Фильтрация информации, полученной из csv файла в нужный формат
            Отбор нужных вакансий, отсев html тегов и т.д.
               :return: void
        '''
        self.filterVacancies = []
        for line in self.otherLines:
            newDict = dict(zip(self.firstLine, line))
            newDict["is_needed"] = newDict["name"].find(self.inputValues.professionName) > -1
            # vacancy = Vacancy(newDict)
            self.filterVacancies.append(Vacancy(newDict))
        self.necessaryVacancies = list(filter(lambda v: v.is_needed, self.filterVacancies))

    def getYears(self):
        '''
        Установка поля allYears после фильтрации вакансий
            :return: void
        '''
        self.allYears = list(set([vacancies.dic["year"] for vacancies in self.filterVacancies]))
        self.allYears.sort()

    def getAverageSalary(self, count: dict, sum: dict) -> dict:
        '''
        Получение средней зарплаты
            :param count (dict): Словарь с количеством вакансий по годам
            :param sum (dict): Словарь с суммой зп по годам
            :return: dict
            >>> DataSet().getAverageSalary(count={'2007' : 2}, sum={'2007' : 10})
            {'2007': 5}
        '''
        keySalary = {}
        for key, value in count.items():
            if value == 0:
                keySalary[key] = 0
            else:
                keySalary[key] = math.floor(sum[key] / value)
        return keySalary

    def getSortedDict(self, keySalary: dict):
        '''
        Сортировка словаря, обрезка его количества до 10
            :param keySalary:
            :return: dict
            >>> len(DataSet().getSortedDict(keySalary={'1': 1, '2' : 2, '3' : 3, '4' : 4, '5' : 5, '6' : 6, '7' : 7, '8' : 8, '9' : 9, '10' : 10, '11' : 11, '12' : 12}))
            10
        '''
        return dict(list(sorted(keySalary.items(), key=lambda item: item[1], reverse=True))[:10])

    def updateKeys(self, keyCount: dict) -> dict:
        '''
            Обновление данных в словаре, если данных там нет
            :param keyCount: Количество вакансий по годам
            :return: dict
        >>> DataSet().updateKeys(keyCount={'2022' : 1000})
        {'2022': 1000, 2007: 0, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}
        >>> DataSet().updateKeys(keyCount={'2023' : 1000})
        {'2023': 1000, 2007: 0, 2008: 0, 2009: 0, 2010: 0, 2011: 0, 2012: 0, 2013: 0, 2014: 0, 2015: 0, 2016: 0, 2017: 0, 2018: 0, 2019: 0, 2020: 0, 2021: 0, 2022: 0}
        '''

        for key in self.allYears:
            if key not in keyCount.keys():
                keyCount[key] = 0
        return keyCount

    def getKey(self, vacancies: list, key_str: str, isArea: bool) -> (dict, dict):
        '''
            Получение данных о зп и кол-ве вакансий по ключу key_str
            :param vacancies: Все вакансии
            :param key_str: Ключ для получения данных
            :param isArea: Флаг
            :return: (dict, dict)
        '''
        keySum = {}
        keyCount = {}
        for v in vacancies:
            keySum = self.tryToAdd(keySum, v.dic[key_str], v.salary.salaryRub)
            keyCount = self.tryToAdd(keyCount, v.dic[key_str], 1)
        if not isArea:
            keySum = self.updateKeys(keySum)
            keyCount = self.updateKeys(keyCount)
        else:
            keyCount = dict(filter(lambda item: item[1] / len(vacancies) > 0.01, keyCount.items()))
        keyAverageSalary = self.getAverageSalary(keyCount, keySum)
        return keyAverageSalary, keyCount

    def numberGraph(self):
        '''
            Установка поля для вывода
            :return: void
        '''
        count_vacs = len(self.filterVacancies)
        self.yearSalary, self.year_to_count = \
            self.getKey(self.filterVacancies, "year", False)
        self.yearSalary_needed, self.yearCount = \
            self.getKey(self.necessaryVacancies, "year", False)
        self.areaSalary, self.area_to_count = \
            self.getKey(self.filterVacancies, "area_name", True)
        self.areaSalary = self.getSortedDict(self.areaSalary)
        self.areaPiece = {key: round(val / count_vacs, 4) for key, val in self.area_to_count.items()}
        self.areaPiece = self.getSortedDict(self.areaPiece)

    def printGraph(self):
        '''
            Вывод полученной аналитики в консоль
            :return: void
        '''
        print("Динамика уровня зарплат по годам:", self.yearSalary)
        print("Динамика количества вакансий по годам:", self.year_to_count)
        print("Динамика уровня зарплат по годам для выбранной профессии:", self.yearSalary_needed)
        print("Динамика количества вакансий по годам для выбранной профессии:", self.yearCount)
        print("Уровень зарплат по городам (в порядке убывания):", self.areaSalary)
        print("Доля вакансий по городам (в порядке убывания):", self.areaPiece)
        Report(inputValues=self.inputValues, yearSalary=self.yearSalary, yearSalary_needed=self.yearSalary_needed,
               year_to_count=self.year_to_count, yearCount=self.yearCount, areaSalary=self.areaSalary,
               areaPiece=self.areaPiece)

class Report:
    '''
        Класс для создания отчетов

        Attributes:
            inputValues (InputConect) Входные данные
            yearSalary (dict) Зарплата по годам
            yearSalary_needed (dict) Зарплата по  годам для вакансии
            year_to_count (dict) Ваканси по годам
            yearCount (dict) Вакансии по годам для профессии
            areaSalary (dict) Уровень зп по городам
            areaPiece (dict) Доля вакансий по городам
        :return: void
    '''

    def __init__(self, inputValues, yearSalary, yearSalary_needed, year_to_count, yearCount, areaSalary, areaPiece):
        self.inputValues = inputValues
        self.yearSalary = yearSalary
        self.yearSalary_needed = yearSalary_needed
        self.year_to_count = year_to_count
        self.yearCount = yearCount
        self.areaSalary = areaSalary
        self.areaPiece = areaPiece
        if self.inputValues.TableOrPdf == "Вакансии":
            self.generate_excel()
        if self.inputValues.TableOrPdf == "Статистика":
            self.generate_excel()
            self.generate_image()
            self.generate_pdf()

    def generate_excel(self):
        '''
            Создание файла отчета в экселе
            :return: void
        '''
        execelFile = Workbook()
        execelSheetFirst = execelFile.create_sheet(title="Статистика по годам", index=0)
        execelSheetFirst['A1'] = 'Год'
        execelSheetFirst['B1'] = 'Средняя зарплата'
        execelSheetFirst['C1'] = f'Средняя зарплата - {self.inputValues.professionName}'
        execelSheetFirst['D1'] = 'Количество вакансий'
        execelSheetFirst['E1'] = f'Количество вакансий - {self.inputValues.professionName}'
        execelSheetFirst['A1'].font = Font(bold=True)
        execelSheetFirst['B1'].font = Font(bold=True)
        execelSheetFirst['C1'].font = Font(bold=True)
        execelSheetFirst['D1'].font = Font(bold=True)
        execelSheetFirst['E1'].font = Font(bold=True)
        yearRow = list(self.yearSalary.keys())
        firstValue = list(self.yearSalary.values())
        secondValues = list(self.yearSalary_needed.values())
        thirdValues = list(self.year_to_count.values())
        fourthValues = list(self.yearCount.values())
        for i in range(0, 16):
            data = list()
            data.append(yearRow[i])
            data.append(firstValue[i])
            data.append(secondValues[i])
            data.append(thirdValues[i])
            data.append(fourthValues[i])
            execelSheetFirst.append(data)

        self.setBorder(columns=['A', 'B', 'C', 'D', 'E'], excelSheet=execelSheetFirst, numberSheet=0)
        self.setSizeColumns(excelSheet=execelSheetFirst)

        execelSheetSecond = execelFile.create_sheet(title="Статистика по городам", index=1)

        execelSheetSecond['A1'] = 'Город'
        execelSheetSecond['B1'] = 'Уровень зарплат'
        execelSheetSecond['D1'] = 'Город'
        execelSheetSecond['E1'] = 'Доля вакансий'
        execelSheetSecond['A1'].font = Font(bold=True)
        execelSheetSecond['B1'].font = Font(bold=True)
        execelSheetSecond['D1'].font = Font(bold=True)
        execelSheetSecond['E1'].font = Font(bold=True)
        cityes1 = list(self.areaSalary.keys())
        salaryes = list(self.areaSalary.values())
        cityes2 = list(self.areaPiece.keys())
        pieces = list(self.areaPiece.values())
        for i in range(0, 10):
            data = list()
            data.append(cityes1[i])
            data.append(salaryes[i])
            data.append("")
            data.append(cityes2[i])
            data.append(pieces[i])
            execelSheetSecond.append(data)

        self.setBorder(columns=['A', 'B', 'D', 'E'], excelSheet=execelSheetSecond, numberSheet=1)
        self.setSizeColumns(excelSheet=execelSheetSecond)

        execelFile.save('report.xlsx')

    def setSizeColumns(self, excelSheet):
        '''
            Установка размера колонок
            :param excelSheet: Экземпляр листа эксель
            :return: void
        '''
        i = 0
        col_width = list()
        for col in excelSheet.columns:
            for j in range(len(col)):
                if j == 0:
                    col_width.append(len(str(col[j].value)))
                else:
                    if col_width[i] < len(str(col[j].value)):
                        col_width[i] = len(str(col[j].value))
            i = i + 1

        for i in range(len(col_width)):
            col_letter = get_column_letter(i + 1)
            if col_width[i] > 100:
                excelSheet.column_dimensions[col_letter].width = 100
            else:
                excelSheet.column_dimensions[col_letter].width = col_width[i] + 2

    def setBorder(self, columns, excelSheet, numberSheet):
        '''
            Установка рамки в экселе
            :param columns: Колонка в экселе
            :param excelSheet: Экземпляр листа в экселе
            :param numberSheet: Номер листа
            :return: void
        '''
        side = Side(border_style='thin', color="00000000")
        for i in columns:
            column = excelSheet[i]
            for j in column:
                j.border = Border(left=side, right=side, top=side, bottom=side)
                if i == 'E' and numberSheet == 1:
                    j.number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

    def generate_image(self):
        '''Создание графиков для аналитики c помощью matplotlib
            :return: void
        '''
        years = [i for i in self.yearSalary.keys()]
        yearSalary = [i for i in self.yearSalary_needed.values()]
        avg = [i for i in self.yearSalary.values()]
        fig = plt.figure()
        ax = fig.add_subplot(221)
        plt.rcParams.update({'font.size': 8})
        ax.set_title("Уровень зарплат по годам", fontsize=8)
        x_nums = range(len(self.yearSalary.keys()))
        width = 0.4
        x_list1 = list(map(lambda x: x - width / 2, x_nums))
        x_list2 = list(map(lambda x: x + width / 2, x_nums))
        ax.bar(x_list1, yearSalary, width, label="средняя з/п")
        ax.bar(x_list2, avg, width, label="з/п аналитик")
        ax.set_xticks(x_nums, years, rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам", fontsize=8)
        vacancies = [i for i in self.yearCount.values()]
        vacanciesAvg = [i for i in self.year_to_count.values()]
        ax.bar(x_list1, vacanciesAvg, width, label="Количество вакансий")
        ax.bar(x_list2, vacancies, width, label="Количество вакансий аналитик")
        ax.set_xticks(x_nums, years, rotation="vertical")
        ax.tick_params(axis="both", labelsize=8)
        ax.legend(fontsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(223)
        ax.invert_yaxis()
        # ax.invert_xaxis()
        ax.set_title("Уровень зарплат по городам", fontsize=8)
        cyties = [i for i in self.areaSalary.keys()]
        salary = [i for i in self.areaSalary.values()]
        ax.barh(cyties, salary, width, label="уровень з/п", align='center')
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        ax.set_title("Доля вакансий по городам", fontsize=8)
        cyties = [i for i in self.areaPiece.keys()]
        values = [i for i in self.areaPiece.values()]
        ax.pie(values, labels=cyties)
        ax.tick_params(axis="both", labelsize=8)

        plt.tight_layout()
        plt.savefig("graph.png")
        plt.show()

    def generate_pdf(self):
        '''Генерация pdf отчета
            :return: void
        '''
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template('pdf_template.html')
        heads1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.inputValues.professionName}',
                  'Количество вакансий', f'Количество вакансий - {self.inputValues.professionName}']
        heads2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
        for key in self.areaPiece.keys():
            self.areaPiece[key] = str(round(self.areaPiece[key] * 100, 2)) + '%'
        pdf_template = template.render({"yearSalary": self.yearSalary,
                                        "yearSalary_needed": self.yearSalary_needed,
                                        "year_to_count": self.year_to_count,
                                        "yearCount": self.yearCount,
                                        "areaSalary": self.areaSalary,
                                        "areaPiece": self.areaPiece,
                                        "heads1": heads1,
                                        "heads2": heads2, })
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": None})
        # for key1, value1, key2, value2 in self.areaSalary.items(), self.areaPiece.items():
        #     print(key1)

class InputConect:
    '''
        Класс для получения входных данных. А также их валидации
        Attributes:
            fileName (string) имя файла для получения статистики
            professionName (string) название профессии, по которой нужна статистика
            TableOrPdf (string) В какои формате нужна статистика
        :return: void
    '''

    def __init__(self):
        self.fileName = 'vacancies_by_year.csv'
        self.professionName = 'Аналитик'
        self.TableOrPdf = 'Вакансии'
        self.checkFile()

    def checkFile(self):
        '''
        Валидация полей
            :return: void
        '''
        with open(self.fileName, "r", encoding='utf-8-sig', newline='') as csv_file:
            file_iter = iter(csv.reader(csv_file))
            if next(file_iter, "none") == "none":
                print("Пустой файл")
                sys.exit()

            if next(file_iter, "none") == "none":
                print("Нет данных")
                sys.exit()

'''Запуск'''
if __name__ == "__main__":
    import doctest

    doctest.testmod()
DataSet()